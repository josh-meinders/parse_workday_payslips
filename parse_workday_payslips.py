import os
from openpyxl import load_workbook
import csv

FIELD_NAMES = ('Company Information - Name',
               'Company Information - Address',
               'Company Information - Phone',
               'Payslip Information - Name',
               'Payslip Information - Employee ID',
               'Payslip Information - Pay Period Begin',
               'Payslip Information - Pay Period End',
               'Payslip Information - Check Date',
               'Payslip Information - Check Number',
               'Current and YTD Totals - Current - Gross Pay',
               'Current and YTD Totals - Current - Pre Tax Deductions',
               'Current and YTD Totals - Current - Employee Taxes',
               'Current and YTD Totals - Current - Post Tax Deductions',
               'Current and YTD Totals - Current - Net Pay',
               'Current and YTD Totals - YTD - Gross Pay',
               'Current and YTD Totals - YTD - Pre Tax Deductions',
               'Current and YTD Totals - YTD - Employee Taxes',
               'Current and YTD Totals - YTD - Post Tax Deductions',
               'Current and YTD Totals - YTD - Net Pay',
               'Earnings - Bonus - Dates',
               'Earnings - Bonus - Hours',
               'Earnings - Bonus - Rate',
               'Earnings - Bonus - Amount',
               'Earnings - Bonus - YTD',
               'Earnings - Bonus Wages - Gross Up - Dates',
               'Earnings - Bonus Wages - Gross Up - Hours',
               'Earnings - Bonus Wages - Gross Up - Rate',
               'Earnings - Bonus Wages - Gross Up - Amount',
               'Earnings - Bonus Wages - Gross Up - YTD',
               'Earnings - Imputed Income Life - Dates',
               'Earnings - Imputed Income Life - Hours',
               'Earnings - Imputed Income Life - Rate',
               'Earnings - Imputed Income Life - Amount',
               'Earnings - Imputed Income Life - YTD',
               'Earnings - Regular Wages - Dates',
               'Earnings - Regular Wages - Hours',
               'Earnings - Regular Wages - Rate',
               'Earnings - Regular Wages - Amount',
               'Earnings - Regular Wages - YTD',
               'Earnings - Spouse/Dep Life Ins Imputed Income - Dates',
               'Earnings - Spouse/Dep Life Ins Imputed Income - Hours',
               'Earnings - Spouse/Dep Life Ins Imputed Income - Rate',
               'Earnings - Spouse/Dep Life Ins Imputed Income - Amount',
               'Earnings - Spouse/Dep Life Ins Imputed Income - YTD',
               'Employee Taxes - OASDI - Amount',
               'Employee Taxes - OASDI - YTD',
               'Employee Taxes - Medicare - Amount',
               'Employee Taxes - Medicare - YTD',
               'Employee Taxes - Federal Withholding - Amount',
               'Employee Taxes - Federal Withholding - YTD',
               'Employee Taxes - State Tax - IA - Amount',
               'Employee Taxes - State Tax - IA - YTD',
               'Pre Tax Deductions - Dental Insurance - Amount',
               'Pre Tax Deductions - Dental Insurance - YTD',
               'Pre Tax Deductions - Health Savings Account - EE - Amount',
               'Pre Tax Deductions - Health Savings Account - EE - YTD',
               'Post Tax Deductions - 403(b) Contribution - Roth - Amount',
               'Post Tax Deductions - 403(b) Contribution - Roth - YTD',
               'Employer Paid Benefits - Basic Retirement Contribution - Amount',
               'Employer Paid Benefits - Basic Retirement Contribution - YTD',
               'Employer Paid Benefits - Disability & Long Term Care - Amount',
               'Employer Paid Benefits - Disability & Long Term Care - YTD',
               'Employer Paid Benefits - Employee Basic Life Premium - Amount',
               'Employer Paid Benefits - Employee Basic Life Premium - YTD',
               'Employer Paid Benefits - Health Savings Account - ER - Amount',
               'Employer Paid Benefits - Health Savings Account - ER - YTD',
               'Employer Paid Benefits - Medical - ER - Amount',
               'Employer Paid Benefits - Medical - ER - YTD',
               'Employer Paid Benefits - Medical Subsidy - Amount',
               'Employer Paid Benefits - Medical Subsidy - YTD',
               'Employer Paid Benefits - Medical Subsidy (Reversal) - Amount',
               'Employer Paid Benefits - Medical Subsidy (Reversal) - YTD',
               'Employer Paid Benefits - Spouse Basic Life - Amount',
               'Employer Paid Benefits - Spouse Basic Life - YTD',
               'Employer Paid Benefits - Worker\'s Comp - US - Amount',
               'Employer Paid Benefits - Worker\'s Comp - US - YTD',
               'Taxable Wages - OASDI - Taxable Wages - Amount',
               'Taxable Wages - OASDI - Taxable Wages - YTD',
               'Taxable Wages - Medicare - Taxable Wages - Amount',
               'Taxable Wages - Medicare - Taxable Wages - YTD',
               'Taxable Wages - Federal Withholding - Taxable Wages - Amount',
               'Taxable Wages - Federal Withholding - Taxable Wages - YTD',
               'Withholding - Marital Status - Federal',
               'Withholding - Marital Status - Work State',
               'Withholding - Allowances - Federal',
               'Withholding - Allowances - Work State',
               'Withholding - Additional Withholding - Federal',
               'Withholding - Additional Withholding - Work State',
               'Payment Information - Bank',
               'Payment Information - Account Name',
               'Payment Information - Account Number',
               'Payment Information - Amount in Pay Group Currency',
               'Payment Information - Pay Group Currency')


def parse_payslip(filename):

    def parse_table_single_row(start_row, payslip_data):
        key_prefix = ws.cell(start_row, 1).value
        label_row = start_row + 1
        data_row = start_row + 2
        col = 1
        while (ws.cell(label_row, col).value != None):
            key_suffix = ws.cell(label_row, col).value
            key = f'{key_prefix} - {key_suffix}'
            val = ws.cell(data_row, col).value
            payslip_data[key] = val
            col += 1

    def parse_table_grid(start_row, payslip_data):
        key_prefix = ws.cell(start_row, 1).value
        label_row = start_row + 1
        data_row = start_row + 2
        while (ws.cell(data_row, 2).value != None):
            col = 2
            key_suffix_1 = ws.cell(data_row, 1).value
            while (ws.cell(label_row, col).value != None):
                key_suffix_2 = ws.cell(label_row, col).value
                val = ws.cell(data_row, col).value
                key = f'{key_prefix} - {key_suffix_1} - {key_suffix_2}'
                payslip_data[key] = val
                col += 1
            data_row += 1

    payslip_data = {}
    ws = load_workbook(filename).active
    for cell in ws['A']:
        if cell.value == 'Company Information':
            parse_table_single_row(cell.row, payslip_data)
        if cell.value == 'Payslip Information':
            parse_table_single_row(cell.row, payslip_data)
        if cell.value == 'Current and YTD Totals':
            parse_table_grid(cell.row, payslip_data)
        if cell.value == 'Earnings':
            parse_table_grid(cell.row, payslip_data)
        if cell.value == 'Employee Taxes':
            parse_table_grid(cell.row, payslip_data)
        if cell.value == 'Pre Tax Deductions':
            parse_table_grid(cell.row, payslip_data)
        if cell.value == 'Post Tax Deductions':
            parse_table_grid(cell.row, payslip_data)
        if cell.value == 'Employer Paid Benefits':
            parse_table_grid(cell.row, payslip_data)
        if cell.value == 'Taxable Wages':
            parse_table_grid(cell.row, payslip_data)
        if cell.value == 'Withholding':
            parse_table_grid(cell.row, payslip_data)
        if cell.value == 'Payment Information':
            parse_table_single_row(cell.row, payslip_data)

    return payslip_data


if __name__ == '__main__':

    with open('output.csv', 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_NAMES)
        writer.writeheader()

        for filename in os.listdir(os.getcwd()):
            if(filename[-5:] == '.xlsx'):
                print(f'Parsing {filename}...')
                writer.writerow(parse_payslip(filename))
